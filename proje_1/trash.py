import random
import os
import numpy  as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
from data import random_numbers, rounded_floats

def file_open_creat(dosya_adi):
  if os.path.exists(dosya_adi):
    #Mevcut dosyayi aç ve değişkene ata
    woorkBook = load_workbook(dosya_adi)
  else:
    # Workbook oluştur
    woorkBook = Workbook()
    
    #worksheet = woorkBook.active # worksheet değişkenine excel çalişma kitabinda ki aktif sayfayi atar
    sayfa_1 = woorkBook.create_sheet("Sayfa 1") #Sayfa 1 adli bir sayfa oluştur
    woorkBook.active = sayfa_1 #Sayfa 1 adli sayfayi aktif sayfa yap

    default_sheet = woorkBook["Sheet"] #Sheet adli sayfayi seçip default_sheet değişkenine ata
    woorkBook.remove(default_sheet) #default_sheet değikenine atanan sayfayi sil 

    sayfa_1.cell(row=1, column=1, value="Miktar") #ilk kolonun adini Miktar olarak belirle
    sayfa_1.cell(row=1, column=2, value="Oran") #ikinci kolonun adini Oran olarak belirle

    # Verileri satir satir yaz
    for i, (deger1, deger2) in enumerate(zip(random_numbers, rounded_floats), start=2):  # start=2 çünkü veriler ikinci satırdan sonra başlıyor. İlk satır başlık olduğu için.

      # Integer değerler için binlik ayraçli format
      column1 = sayfa_1.cell(row=i, column=1, value=deger1)
      column1.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2  # Binlik ayraçli ve 2 ondalik basamak

      # Float değerler için binlik ayraçli ve virgülden sonra 2 basamakli format
      column2 = sayfa_1.cell(row=i, column=2, value=deger2)
      column2.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2  # Binlik ayraçli ve 2 ondalik basamak

  return woorkBook

"""
zip(random_numbers, rounded_floats)
zip() fonksiyonu, birden fazla iterable'i (liste, tuple, vb.) birleştirir ve her bir iterable'dan birer eleman alarak bir tuple oluşturur.
[(1, 1.1), (2, 2.2), (3, 3.3)] gibi
"""

"""
enumerate() fonksiyonu, bir iterable'in her bir elemanini indeksler. Varsayilan olarak indeksler 0'dan başlar, ancak start=2 parametresi ile indekslemeyi 2'den başlatir.

Örneğin, zip(random_numbers, rounded_floats) sonucu [(1, 1.1), (2, 2.2), (3, 3.3)] ise, enumerate() şu şekilde bir iterable oluşturur:
[(2, (1, 1.1)), (3, (2, 2.2)), (4, (3, 3.3))]
"""

"""
for i, (deger1, deger2) in ... >> Bu kisim, enumerate() ile oluşturulan iterable üzerinde döngü yapar. i, indeks değerini temsil eder (2'den başlayarak).

(deger1, deger2), zip() ile oluşturulan tuple'i ayriştirir. deger1, random_numbers listesinden bir eleman, deger2 ise rounded_floats listesinden bir elemandir.
"""
dosya = "work_datas.xlsx"

woorkBook = file_open_creat(dosya)

sayfa_1 = woorkBook["Sayfa 1"]

sayfa_1.cell(row=1, column=3, value="deneme başlik")

woorkBook.save(dosya)


"""
print(random.random())

#random.randint(a, b) >> Açiklama: a ile b arasinda (dahil) rastgele bir tam sayi (integer) döndürür.
print(random.randint(1, 10))

#random.uniform(a, b) >> Açiklama: a ile b arasinda rastgele bir kayan noktali sayi (float) döndürür.
print(random.uniform(1, 5))

#random.choice(seq) >> Açiklama: Bir diziden (liste, tuple, string vb.) rastgele bir öğe seçer.
my_list = ["elma", "armut", "çilek"]
print(random.choice(my_list))

#random.shuffle(seq) >> Açiklama: Bir dizinin (liste) öğelerini rastgele kariştirir. Listenin kendisini değiştirir.
my_list = [1, 2, 3, 4, 5]
random.shuffle(my_list)
print(my_list)

#random.sample(population, k) >> Açiklama: Bir diziden (population) belirli sayida (k) benzersiz öğe seçer.
my_list = [1, 2, 3, 4, 5]
print(random.sample(my_list, 3))

#random.seed(a=None) >> Açiklama: Rastgele sayi üretecinin başlangiç değerini (seed) ayarlar. Ayni seed değeri kullanildiğinda, ayni rastgele sayilar üretilir. Bu, tekrarlanabilirlik sağlar.
#random.seed(42)  # Seed değerini ayarla
print(random.randint(1, 100))  # Her zaman ayni sonucu verir, örneğin: 82

#random.randrange(start, stop, step) >> Açiklama: Belirli bir aralikta (start ile stop arasinda) ve belirli bir adimla (step) rastgele bir tam sayi döndürür.
print(random.randrange(0, 10, 2))  # Örnek çikti: 6 (0, 2, 4, 6, 8 arasindan)

#random.choices(population, weights=None, k=1) >> Açiklama: Bir diziden (population) belirli sayida (k) öğe seçer. weights parametresi ile her öğenin seçilme olasiliğini ayarlayabilirsiniz.
my_list = ["elma", "armut", "çilek"]
print(random.choices(my_list, weights=[10, 1, 1], k=2))  # Örnek çikti: ["elma", "elma"]

#random.gauss(mu, sigma) >> Açiklama: Normal dağilima (Gauss dağilimi) uygun bir rastgele sayi döndürür. mu ortalama, sigma standart sapmadir.
print(random.gauss(0, 1))  # Örnek çikti: 0.123456 (ortalama 0, standart sapma 1)

#random.getrandbits(k) >>Açiklama: k bit uzunluğunda rastgele bir tam sayi döndürür.
print(random.getrandbits(8))  # Örnek çikti: 173 (0 ile 255 arasinda bir sayi)

#random.triangular(low, high, mode) >> Açiklama: Üçgen dağilima uygun bir rastgele sayi döndürür. low minimum, high maksimum, mode en olasi değerdir.
print(random.triangular(0, 10, 5))  # Örnek çikti: 4.567

#random.betavariate(alpha, beta) >> Açiklama: Beta dağilimina uygun bir rastgele sayi döndürür.
print(random.betavariate(0.5, 0.5))  # Örnek çikti: 0.123456

#random.expovariate(lambd) >> Açiklama: Üstel dağilima uygun bir rastgele sayi döndürür. lambd, dağilimin parametresidir.
print(random.expovariate(1.5))  # Örnek çikti: 0.456789

#random.gammavariate(alpha, beta) >> Açiklama: Gamma dağilimina uygun bir rastgele sayi döndürür.
print(random.gammavariate(1.0, 2.0))  # Örnek çikti: 1.234567
"""