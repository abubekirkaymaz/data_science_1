import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers
from data import random_numbers, rounded_floats


def open_creat_file (file): 
  if os.path.exists(file):
    workBook = load_workbook(file)

  else: 
    workBook = Workbook()

    sayfa1 = workBook.create_sheet("Sayfa 1")
    workBook.active = sayfa1
    default_sheet = workBook["Sheet"]
    workBook.remove(default_sheet)
    

    sayfa1.cell(row=1, column=1, value = "Miktar")
    sayfa1.cell(row=1, column=2, value = "Oran")

    for indeks, (dizi1, dizi2) in enumerate(zip(random_numbers, rounded_floats), start=2):

      column1 = sayfa1.cell(row=indeks, column=1, value = dizi1)
      column1.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

      column2 = sayfa1.cell(row=indeks, column=2, value = dizi2)
      column2.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

  return workBook


dosya = "work_datas.xlsx"

workBook = open_creat_file(dosya)

sayfa_1 = workBook["Sayfa 1"]
sayfa_1.cell(row=1, column=3, value="Deneme")

workBook.save(dosya)

print(random_numbers)
print(rounded_floats)


