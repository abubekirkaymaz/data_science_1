from openpyxl import load_workbook

def calculate_kdv(amount, rate):
  vat_amount = amount * rate / 100
  gross_amount = amount + vat_amount
  return vat_amount, gross_amount

# Excel dosyasını yükle
workbook = load_workbook("veriler.xlsx")
worksheet = workbook["Sayfa1"]  # "Sayfa1" yerine kendi sayfa adınızı yazın

# Başlıkları ekle (isteğe bağlı)
worksheet.cell(row=1, column=3).value = "KDV Tutarı"
worksheet.cell(row=1, column=4).value = "Brüt Tutar"

# Verileri oku, işle ve yaz
for i, row in enumerate(worksheet.iter_rows(min_row=2)):  # min_row=2 ile başlığı atla
  net_fiyat = row[0].value
  kdv_orani = row[1].value

  if isinstance(net_fiyat, (int, float)) and isinstance(kdv_orani, (int, float)): #Veri kontrolü eklendi.
    kdv_tutari, brut_tutar = calculate_kdv(net_fiyat, kdv_orani)

    worksheet.cell(row=i + 2, column=3).value = kdv_tutari
    worksheet.cell(row=i + 2, column=4).value = brut_tutar
  else:
    print(f"Hata: Geçersiz veri türü. Satır {i+2}") #Hata mesajı eklendi.

# Excel dosyasını kaydet (aynı dosyaya)
workbook.save("veriler.xlsx")

print("İşlem tamamlandı.")

